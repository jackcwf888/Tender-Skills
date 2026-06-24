import argparse
import asyncio
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
import edge_tts


DEFAULT_VOICE = "zh-CN-YunyangNeural"
DEFAULT_RATE = "+0%"
DEFAULT_PITCH = "-1Hz"
DEFAULT_VOLUME = "+0%"
DEFAULT_PAUSE_MS = 650


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate tender-style Chinese roadshow voiceover MP3 from Excel or text."
    )
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--xlsx", help="Path to source Excel workbook")
    source.add_argument("--text-file", help="Path to plain-text narration file")
    parser.add_argument("--sheet", default="路演脚本拆解", help="Worksheet name for Excel input")
    parser.add_argument(
        "--text-column",
        default="旁白文本",
        help="Column header containing narration text in row 1",
    )
    parser.add_argument("--start-row", type=int, default=2, help="First data row for Excel input")
    parser.add_argument("--end-row", type=int, help="Last data row for Excel input")
    parser.add_argument("--voice", default=DEFAULT_VOICE, help="Edge TTS voice")
    parser.add_argument("--rate", default=DEFAULT_RATE, help="Speaking rate, e.g. +0%, -15%")
    parser.add_argument("--pitch", default=DEFAULT_PITCH, help="Pitch, e.g. -1Hz")
    parser.add_argument("--volume", default=DEFAULT_VOLUME, help="Volume, e.g. +0%")
    parser.add_argument("--pause-ms", type=int, default=DEFAULT_PAUSE_MS, help="Pause between lines")
    parser.add_argument("--ffmpeg", default="ffmpeg", help="Path to ffmpeg")
    parser.add_argument("--ffprobe", default="ffprobe", help="Path to ffprobe")
    parser.add_argument("--output", required=True, help="Final MP3 output path")
    parser.add_argument("--keep-temp", action="store_true", help="Keep temp synthesis directory")
    return parser.parse_args()


def normalize_text(text: str) -> str:
    return str(text).replace("\r", "").replace("\n", "，").strip()


def extract_lines_from_xlsx(path: Path, sheet_name: str, text_column: str, start_row: int, end_row: int | None) -> list[str]:
    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {sheet_name}")
    sheet = workbook[sheet_name]

    header_map = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=col).value
        if value is not None:
            header_map[str(value).strip()] = col
    if text_column not in header_map:
        raise ValueError(f"Text column not found in row 1: {text_column}")

    text_col = header_map[text_column]
    last_row = end_row or sheet.max_row
    lines = []
    for row in range(start_row, last_row + 1):
        value = sheet.cell(row=row, column=text_col).value
        if value is None:
            continue
        normalized = normalize_text(value)
        if normalized:
            lines.append(normalized)
    return lines


def extract_lines_from_text(path: Path) -> list[str]:
    raw = path.read_text(encoding="utf-8")
    chunks = [normalize_text(line) for line in raw.splitlines()]
    return [line for line in chunks if line]


async def synthesize_segments(lines: list[str], segment_dir: Path, voice: str, rate: str, pitch: str, volume: str) -> None:
    for index, line in enumerate(lines, start=1):
        output = segment_dir / f"line_{index:03d}.mp3"
        last_error = None
        for _ in range(3):
            try:
                communicator = edge_tts.Communicate(
                    text=line,
                    voice=voice,
                    rate=rate,
                    pitch=pitch,
                    volume=volume,
                )
                await communicator.save(str(output))
                last_error = None
                break
            except Exception as exc:
                last_error = exc
                await asyncio.sleep(1)
        if last_error is not None:
            raise RuntimeError(f"Failed to synthesize line {index}: {last_error}") from last_error


def run_command(args: list[str]) -> None:
    subprocess.run(args, check=True)


def build_pause_file(ffmpeg: str, pause_ms: int, pause_file: Path) -> None:
    duration = f"{pause_ms / 1000:.3f}"
    run_command(
        [
            ffmpeg,
            "-y",
            "-f",
            "lavfi",
            "-i",
            "anullsrc=r=24000:cl=mono",
            "-t",
            duration,
            "-q:a",
            "9",
            "-acodec",
            "libmp3lame",
            str(pause_file),
        ]
    )


def concat_segments(ffmpeg: str, segment_dir: Path, pause_file: Path, line_count: int, output: Path) -> None:
    concat_file = segment_dir.parent / "concat_list.txt"
    lines = []
    for index in range(1, line_count + 1):
        segment = (segment_dir / f"line_{index:03d}.mp3").as_posix()
        lines.append(f"file '{segment}'")
        if index < line_count:
            lines.append(f"file '{pause_file.as_posix()}'")
    concat_file.write_text("\n".join(lines), encoding="utf-8")
    run_command(
        [
            ffmpeg,
            "-y",
            "-f",
            "concat",
            "-safe",
            "0",
            "-i",
            str(concat_file),
            "-c:a",
            "libmp3lame",
            "-b:a",
            "128k",
            str(output),
        ]
    )


def probe_audio(ffprobe: str, output: Path) -> dict:
    result = subprocess.run(
        [
            ffprobe,
            "-v",
            "error",
            "-show_entries",
            "format=duration,bit_rate:stream=codec_name,sample_rate,channels",
            "-of",
            "json",
            str(output),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout)


def main() -> int:
    args = parse_args()
    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    if shutil.which(args.ffmpeg) is None and not Path(args.ffmpeg).exists():
        raise FileNotFoundError(f"ffmpeg not found: {args.ffmpeg}")
    if shutil.which(args.ffprobe) is None and not Path(args.ffprobe).exists():
        raise FileNotFoundError(f"ffprobe not found: {args.ffprobe}")

    if args.xlsx:
        lines = extract_lines_from_xlsx(
            Path(args.xlsx),
            args.sheet,
            args.text_column,
            args.start_row,
            args.end_row,
        )
    else:
        lines = extract_lines_from_text(Path(args.text_file))

    if not lines:
        raise ValueError("No narration lines found in the source.")

    temp_dir_obj = tempfile.TemporaryDirectory(prefix="tender_voiceover_")
    temp_dir = Path(temp_dir_obj.name)
    segment_dir = temp_dir / "segments"
    segment_dir.mkdir(parents=True, exist_ok=True)
    pause_file = temp_dir / "pause.mp3"
    normalized_text = output.with_suffix(".txt")
    normalized_text.write_text("\n".join(lines), encoding="utf-8")

    try:
        asyncio.run(
            synthesize_segments(
                lines,
                segment_dir,
                args.voice,
                args.rate,
                args.pitch,
                args.volume,
            )
        )
        build_pause_file(args.ffmpeg, args.pause_ms, pause_file)
        concat_segments(args.ffmpeg, segment_dir, pause_file, len(lines), output)
        probe = probe_audio(args.ffprobe, output)
        print(json.dumps({"output": str(output), "line_count": len(lines), "audio": probe}, ensure_ascii=True, indent=2))
        if args.keep_temp:
            print(json.dumps({"temp_dir": str(temp_dir)}, ensure_ascii=True))
            temp_dir_obj.cleanup = lambda: None  # type: ignore[attr-defined]
        return 0
    finally:
        if not args.keep_temp:
            temp_dir_obj.cleanup()


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
